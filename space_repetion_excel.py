import openpyxl
import datetime
import schedule
import time
from tkinter import *
from tkinter import messagebox
from tkinter import ttk  # For table display
from plyer import notification
from threading import Thread

# Define the Excel file
excel_file = "spaced_repetition_checklist_with_date_time.xlsx"

# Initialize the Excel file if it doesn't exist
def setup_excel():
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header row with individual revision completion columns
        ws.append(["Topic", "First Revision", "First Revision Done", "Second Revision", "Second Revision Done",
                   "Third Revision", "Third Revision Done", "Fourth Revision", "Fourth Revision Done"])
        wb.save(excel_file)

# Function to add a new topic with spaced repetition schedule in the Excel file
def add_topic(topic):
    if not topic:
        messagebox.showwarning("Input Error", "Please enter a topic.")
        return

    current_time = datetime.datetime.now()
    intervals = [24, 72, 200, 600]  # Time intervals in hours for spaced repetition
    # Store date and time in "DD/MM/YYYY HH:MM" format
    revision_times = [(current_time + datetime.timedelta(hours=hours)).strftime("%d/%m/%Y %H:%M") for hours in intervals]

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Insert the topic and the revision times into the Excel file, all "Done" columns set to "No"
    ws.append([topic, revision_times[0], "No", revision_times[1], "No", revision_times[2], "No", revision_times[3], "No"])

    wb.save(excel_file)
    topic_entry.delete(0, END)  # Clear the text field
    messagebox.showinfo("Success", f"Topic '{topic}' added with revision schedule.")
    display_upcoming_revisions()  # Update table view

# Function to check for topics due for revision from the Excel file
def check_reminders():
    current_time = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")  # Current time in "DD/MM/YYYY HH:MM"
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        topic, first_revision, first_done, second_revision, second_done, \
        third_revision, third_done, fourth_revision, fourth_done = row

        # Check each revision time and send a notification if the time is past or due, and revision is not marked as "Done"
        if first_done == "No" and first_revision and first_revision <= current_time:
            send_notification(f"{topic} (First Revision)")
        if second_done == "No" and second_revision and second_revision <= current_time:
            send_notification(f"{topic} (Second Revision)")
        if third_done == "No" and third_revision and third_revision <= current_time:
            send_notification(f"{topic} (Third Revision)")
        if fourth_done == "No" and fourth_revision and fourth_revision <= current_time:
            send_notification(f"{topic} (Fourth Revision)")

    wb.save(excel_file)

# Function to send a desktop notification
def send_notification(topic):
    notification.notify(
        title="Revision Reminder",
        message=f"Time to revise: {topic}",
        timeout=10  # Notification timeout in seconds
    )

# Function to display upcoming revisions from the Excel file in the GUI
def display_upcoming_revisions():
    # Clear the table first
    for item in revision_table.get_children():
        revision_table.delete(item)

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # Insert rows into the table
    for row in ws.iter_rows(min_row=2, values_only=True):
        revision_table.insert("", "end", values=row)

# Function to handle the scheduled reminders
def run_scheduler():
    while True:
        schedule.run_pending()
        time.sleep(1)

# Create the Tkinter GUI
def create_gui():
    global topic_entry, revision_table

    root = Tk()
    root.title("Spaced Repetition System")
    root.geometry("1050x400")

    # Label for entering a new topic
    Label(root, text="Enter a new topic:").pack(pady=10)

    # Entry box for the topic
    topic_entry = Entry(root, width=50)
    topic_entry.pack(pady=5)

    # Button to add the topic
    add_button = Button(root, text="Add Topic", command=lambda: add_topic(topic_entry.get()))
    add_button.pack(pady=10)

    # Create a table (Treeview) for displaying upcoming revisions
    revision_table = ttk.Treeview(root, columns=("Topic", "First Revision", "First Done", "Second Revision", "Second Done",
                                                 "Third Revision", "Third Done", "Fourth Revision", "Fourth Done"),
                                  show="headings", height=10)
    revision_table.heading("Topic", text="Topic")
    revision_table.heading("First Revision", text="First Revision")
    revision_table.heading("First Done", text="First Done")
    revision_table.heading("Second Revision", text="Second Revision")
    revision_table.heading("Second Done", text="Second Done")
    revision_table.heading("Third Revision", text="Third Revision")
    revision_table.heading("Third Done", text="Third Done")
    revision_table.heading("Fourth Revision", text="Fourth Revision")
    revision_table.heading("Fourth Done", text="Fourth Done")
    
    revision_table.column("Topic", width=150, anchor=CENTER)
    revision_table.column("First Revision", width=150, anchor=CENTER)
    revision_table.column("First Done", width=100, anchor=CENTER)
    revision_table.column("Second Revision", width=150, anchor=CENTER)
    revision_table.column("Second Done", width=100, anchor=CENTER)
    revision_table.column("Third Revision", width=150, anchor=CENTER)
    revision_table.column("Third Done", width=100, anchor=CENTER)
    revision_table.column("Fourth Revision", width=150, anchor=CENTER)
    revision_table.column("Fourth Done", width=100, anchor=CENTER)
    revision_table.pack(pady=10)

    # Button to display upcoming revisions
    show_revisions_button = Button(root, text="Show Upcoming Revisions", command=display_upcoming_revisions)
    show_revisions_button.pack(pady=10)

    # Button to exit the program
    exit_button = Button(root, text="Exit", command=root.quit)
    exit_button.pack(pady=10)

    root.mainloop()

# Main function to set up the Excel file and start the reminder system
if __name__ == "__main__":
    # Setup Excel file
    setup_excel()

    # Schedule the reminder checker to run every minute
    schedule.every(1).minute.do(check_reminders)

    # Start the scheduler in a separate thread
    scheduler_thread = Thread(target=run_scheduler)
    scheduler_thread.daemon = True
    scheduler_thread.start()

    # Start the Tkinter GUI
    create_gui()
